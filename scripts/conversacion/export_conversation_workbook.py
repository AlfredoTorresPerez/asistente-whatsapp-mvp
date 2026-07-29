#!/usr/bin/env python3
"""Exporta la planilla de conversacion a recursos JSON versionables.

El runtime no debe leer XLSX. Este script convierte la planilla fuente en
recursos JSON para produccion y fixtures JSON para pruebas.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_SHEETS = [
    "Resumen",
    "Reservar",
    "Reprogramar",
    "Cancelar",
    "Taxonomia_Intenciones",
    "Ejemplos_Ampliados",
    "Escenarios_Multiturno",
    "Conversaciones_Multiturno",
    "Entidades_Estructuradas",
    "Casos_Negativos",
    "Contexto_Referencias",
    "Cambios_Intencion",
    "Multiples_Reservas",
    "Respuestas_Acciones",
    "Estados_Conversacion",
    "Transiciones_Conversacion",
    "Evaluacion_Dataset",
    "Cobertura_Dataset",
    "Ambiguas_Contexto",
    "Reglas_Modelo",
    "Fuentes",
    "Metodologia",
]

SHEET_ALIASES = {
    "Metodolog\ufffda": "Metodologia",
    "Metodología": "Metodologia",
}

MAIN_EXPORTS = {
    "Taxonomia_Intenciones": "intents.json",
    "Estados_Conversacion": "states.json",
    "Transiciones_Conversacion": "transitions.json",
    "Respuestas_Acciones": "responses.json",
    "Reglas_Modelo": "rules.json",
    "Ambiguas_Contexto": "ambiguous-contexts.json",
}

TEST_EXPORTS = {
    "Reservar": "booking-examples.json",
    "Reprogramar": "reschedule-examples.json",
    "Cancelar": "cancel-examples.json",
    "Ejemplos_Ampliados": "expanded-examples.json",
    "Escenarios_Multiturno": "multiturn-scenarios.json",
    "Conversaciones_Multiturno": "multiturn-turns.json",
    "Entidades_Estructuradas": "structured-entities.json",
    "Casos_Negativos": "negative-cases.json",
    "Contexto_Referencias": "context-references.json",
    "Cambios_Intencion": "intent-switches.json",
    "Multiples_Reservas": "multiple-bookings.json",
    "Evaluacion_Dataset": "evaluation-cases.json",
    "Cobertura_Dataset": "coverage.json",
    "Fuentes": "sources.json",
    "Metodologia": "methodology.json",
}


def normalize_sheet_name(value: str) -> str:
    return SHEET_ALIASES.get(value, value)


def normalize_key(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = text.replace("�", "")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "columna"


def normalize_text(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def read_rows(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    normalized_names = {normalize_sheet_name(name): name for name in workbook.sheetnames}
    missing = [name for name in EXPECTED_SHEETS if name not in normalized_names]
    if missing:
        raise SystemExit(f"Faltan hojas esperadas: {', '.join(missing)}")

    output: dict[str, list[dict[str, Any]]] = {}
    for normalized_name, original_name in normalized_names.items():
        if normalized_name not in EXPECTED_SHEETS:
            continue
        worksheet = workbook[original_name]
        iterator = worksheet.iter_rows(values_only=True)
        raw_header = next(iterator, [])
        header = [normalize_key(cell) for cell in raw_header]
        rows: list[dict[str, Any]] = []
        for row in iterator:
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            item: dict[str, Any] = {}
            for key, value in zip(header, row):
                if not key:
                    continue
                item[key] = normalize_text(value)
            rows.append(item)
        output[normalized_name] = rows
    return output


def validate(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []

    for sheet in EXPECTED_SHEETS:
        if sheet not in rows_by_sheet:
            errors.append(f"No se exporto la hoja {sheet}")

    for sheet, rows in rows_by_sheet.items():
        ids = [str(row.get("id")) for row in rows if row.get("id")]
        duplicated = [item for item, count in Counter(ids).items() if count > 1]
        if duplicated:
            errors.append(f"IDs duplicados en {sheet}: {', '.join(duplicated)}")

    intent_names = {row.get("intencion") for row in rows_by_sheet["Taxonomia_Intenciones"] if row.get("intencion")}
    state_names = {row.get("estado") for row in rows_by_sheet["Estados_Conversacion"] if row.get("estado")}
    state_names.add("CUALQUIERA")

    for row in rows_by_sheet["Respuestas_Acciones"]:
        intent = row.get("intencion")
        if intent and intent not in intent_names:
            errors.append(f"Respuesta referencia intencion inexistente: {intent}")
        state = row.get("estado")
        if state and state not in state_names and state != "CUALQUIERA":
            warnings.append(f"Respuesta referencia estado no declarado: {state}")

    for row in rows_by_sheet["Transiciones_Conversacion"]:
        origin = row.get("estado_origen")
        destination = row.get("estado_destino")
        if origin and origin not in state_names:
            errors.append(f"Transicion con estado origen inexistente: {origin}")
        if destination and destination not in state_names:
            errors.append(f"Transicion con estado destino inexistente: {destination}")

    scenario_turns = Counter(row.get("escenario_id") for row in rows_by_sheet["Conversaciones_Multiturno"])
    for row in rows_by_sheet["Escenarios_Multiturno"]:
        scenario_id = row.get("escenario_id")
        expected = row.get("cantidad_de_turnos")
        if scenario_id is None or expected is None:
            continue
        try:
            expected_count = int(expected)
        except (TypeError, ValueError):
            warnings.append(f"Escenario {scenario_id} tiene cantidad de turnos invalida: {expected}")
            continue
        actual = scenario_turns.get(scenario_id, 0)
        if actual != expected_count:
            errors.append(f"Escenario {scenario_id} declara {expected_count} turnos pero tiene {actual}")

    eval_splits = Counter(row.get("conjunto") for row in rows_by_sheet["Evaluacion_Dataset"])
    for split in ("entrenamiento", "validacion", "validaci\ufffdn", "prueba"):
        if split in eval_splits:
            break
    if not eval_splits:
        errors.append("Evaluacion_Dataset no tiene particiones")

    recalculated_coverage = defaultdict(lambda: {"examples": 0, "turns": 0, "negative": 0, "evaluation": 0})
    for sheet in ("Reservar", "Reprogramar", "Cancelar"):
        for row in rows_by_sheet[sheet]:
            intent = row.get("intencion")
            if intent:
                recalculated_coverage[intent]["examples"] += 1
    for row in rows_by_sheet["Ejemplos_Ampliados"]:
        intent = row.get("intencion_esperada")
        if intent:
            recalculated_coverage[intent]["examples"] += 1
    for row in rows_by_sheet["Conversaciones_Multiturno"]:
        intent = row.get("intencion_detectada")
        if intent:
            recalculated_coverage[intent]["turns"] += 1
    for row in rows_by_sheet["Casos_Negativos"]:
        intent = row.get("intencion_correcta")
        if intent:
            recalculated_coverage[intent]["negative"] += 1
    for row in rows_by_sheet["Evaluacion_Dataset"]:
        intent = row.get("intencion_esperada")
        if intent:
            recalculated_coverage[intent]["evaluation"] += 1

    uncovered = sorted(intent for intent in intent_names if intent not in recalculated_coverage)
    if uncovered:
        warnings.append("Intenciones sin cobertura recalculada: " + ", ".join(uncovered))

    return {
        "errors": errors,
        "warnings": warnings,
        "sheet_counts": {sheet: len(rows) for sheet, rows in rows_by_sheet.items()},
        "evaluation_splits": dict(eval_splits),
        "coverage_recalculated": {
            intent: {
                **counts,
                "total": sum(counts.values()),
            }
            for intent, counts in sorted(recalculated_coverage.items())
        },
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta planilla de conversacion a JSON")
    parser.add_argument(
        "--workbook",
        default="documentacion/conversacion/intenciones_agenda_whatsapp_es419_actualizada.xlsx",
        type=Path,
    )
    parser.add_argument("--main-output", default="backend-java/src/main/resources/conversation", type=Path)
    parser.add_argument("--test-output", default="backend-java/src/test/resources/conversation", type=Path)
    parser.add_argument("--fail-on-warning", action="store_true")
    args = parser.parse_args()

    rows_by_sheet = read_rows(args.workbook)
    report = validate(rows_by_sheet)
    if report["errors"] or (args.fail_on_warning and report["warnings"]):
        for error in report["errors"]:
            print("ERROR:", error)
        for warning in report["warnings"]:
            print("WARNING:", warning)
        raise SystemExit(1)

    metadata = {
        "source": str(args.workbook),
        "sheet_counts": report["sheet_counts"],
        "warnings": report["warnings"],
    }
    for sheet, file_name in MAIN_EXPORTS.items():
        write_json(args.main_output / file_name, {"metadata": metadata, "items": rows_by_sheet[sheet]})
    for sheet, file_name in TEST_EXPORTS.items():
        write_json(args.test_output / file_name, {"metadata": metadata, "items": rows_by_sheet[sheet]})
    write_json(args.test_output / "coverage-recalculated.json", report)
    print(json.dumps({"exported": True, **report}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
